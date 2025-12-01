from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path


def update_excel(prices: dict[str, float], spreadsheet_path: Path) -> None:
    wb = load_workbook(spreadsheet_path)
    sheet = wb["List2"]
    # print(sheet.max_row, sheet.min_row, sheet.max_column, sheet.min_column)

    # Expect product names in column A;
    # write price into column B.
    for row in range(1, sheet.max_row + 1):
        product_name = sheet[f"A{row}"].value.strip().lower()
        if product_name in prices:
            sheet[f"B{row}"] = prices[product_name]
            sheet[f"C{row}"] = datetime.now().strftime("%Y-%m-%d")  # optional timestamp

    wb.save(spreadsheet_path)
    print("Excel updated ✔")